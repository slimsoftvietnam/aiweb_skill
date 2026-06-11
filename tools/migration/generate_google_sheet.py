#!/usr/bin/env python3
"""Tạo file Excel 2 sheet — upload lên Google Sheets."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
INVENTORY = ROOT / "output" / "slimcrm_inventory.json"
OUTPUT = ROOT / "output" / "slimcrm_migration_google_sheet.xlsx"
BASE_NEW = "https://your-domain.com"


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)
    ws.freeze_panes = "A2"


def landing_slug(source_key: str) -> str:
    return "home" if source_key == "index" else source_key


def build_checklist() -> list[tuple]:
    return [
        ("Chuẩn bị", "Đọc skill migration pipeline", ".cursor/skills/aiweb-migrate/SKILL.md", "Đã xong", ""),
        ("Chuẩn bị", "Cấu hình migration_api_key trong SQLite", "tools/migration/config.env", "Đã xong", "Không commit key"),
        ("Chuẩn bị", "pip install dependencies", "pip install -r tools/migration/requirements.txt", "Đã xong", ""),
        ("Recon", "Chạy inventory landing (BFS .html)", "python scrapers/slimcrm_recon.py", "Đã xong", "37 landing"),
        ("Recon", "Bỏ qua blog.slimcrm.vn", "(mặc định, không --include-blog)", "Bỏ qua", "Theo yêu cầu"),
        ("Extract", "Extract manifest pilot", "python scrapers/slimcrm_extract.py --pilot", "Đã xong", "5 trang"),
        ("Extract", "Extract manifest full", "python scrapers/slimcrm_extract.py", "Đã xong", "37 trang, 264 ảnh"),
        ("Import", "Ping Migration API", "python runners/import_manifest.py --ping-only", "Đã xong", "AIWEB_BASE=.../aiweb_core"),
        ("Import", "Dry-run import pilot", "--dry-run --file slimcrm_manifest_pilot_landing.json", "Đã xong", ""),
        ("Import", "Import pilot", "--file slimcrm_manifest_pilot_landing.json", "Đã xong", ""),
        ("Import", "Import full landing", "--file slimcrm_manifest_landing.json", "Đã xong", ""),
        ("Publish", "Publish landing", "--publish --force --file slimcrm_manifest_landing.json", "Đã xong", ""),
        ("Verify", "list_entities = 37", "API action=list_entities", "Đã xong", ""),
        ("Verify", "Kiểm tra URL /{slug}", "VD: /home, /tinhnang", "Đã xong", "Không dùng /p/{slug}"),
        ("Verify", "Ảnh rewrite mig_*", "DevTools Network/Elements", "Đã xong", ""),
        ("Tùy chọn", "Migrate blog.slimcrm.vn", "recon/extract --include-blog", "Bỏ qua", "~911 bài"),
        ("Tùy chọn", "Migrate legacy /blog/*.html", "8 URL sheet Links", "Chưa", ""),
        ("Tùy chọn", "Redirect 301 cũ → mới", "Nginx/Apache", "Chưa", ""),
    ]


def main() -> None:
    inv = json.loads(INVENTORY.read_text(encoding="utf-8"))
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Checklist Agent"
    ws1.append(["STT", "Phase", "Việc Agent cần làm", "Lệnh / File", "Trạng thái", "Ghi chú"])
    for i, row in enumerate(build_checklist(), 1):
        ws1.append([i, *row])
    style_sheet(ws1)

    ws2 = wb.create_sheet("Links migrate")
    ws2.append([
        "STT",
        "Loại",
        "Phạm vi migrate",
        "URL nguồn",
        "Tiêu đề",
        "Slug AI Web",
        "URL đích (mẫu)",
        "Trạng thái",
        "Ghi chú",
    ])

    stt = 0
    for page in inv["sources"]["slimcrm.vn"]["landing_pages"]:
        stt += 1
        slug = landing_slug(page["source_key"])
        ws2.append([
            stt,
            "Landing",
            "Có",
            page["url"],
            page.get("title_guess", ""),
            slug,
            f"{BASE_NEW}/{slug}",
            "Đã xong",
            "Import + publish",
        ])

    for page in inv["sources"]["slimcrm.vn"].get("legacy_blog", []):
        stt += 1
        sk = page["source_key"]
        ws2.append([
            stt,
            "Legacy blog (main)",
            "Tùy chọn",
            page["url"],
            page.get("title_guess", ""),
            sk,
            f"{BASE_NEW}/{sk}",
            "Chưa",
            "Chưa trong batch landing",
        ])

    for extra in (
        ("Blog subdomain", "Không", "https://blog.slimcrm.vn/", "(toàn bộ blog)", "-", "Bỏ qua", "911 URL"),
        ("Help docs", "Không", "https://help.slimcrm.vn/", "(tài liệu)", "-", "Bỏ qua", "GitBook"),
    ):
        stt += 1
        ws2.append([stt, extra[0], extra[1], extra[2], extra[3], extra[4], "-", extra[5], extra[6]])

    style_sheet(ws2)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
